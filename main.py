import datetime
import calendar
import imaplib
import logging
import openpyxl as op
import pytz
import schedule
import sqlite3 as sq
import time
import threading

from dateutil.relativedelta import relativedelta
from io import BytesIO
from PIL import Image
from pytz import timezone
from pyzbar.pyzbar import decode
from telebot import TeleBot, types

import api_dj
import api_glonasssoft
import api_mts
from classes import Number
from configs import (DB, telegram_token, telegram_malashin_id, telegram_my_930_id,
                     telegram_sumbulov_id, telegram_my_id,
                     imap_server_yandex, ya_mary_email_login,
                     ya_mary_email_password, id_telegram_mary, timezone_my,
                     telegram_maks_id, glonasssoft_org_id, telegram_job_id)


logging.basicConfig(
    level=logging.INFO,
    filename="logs.log",
    filemode="a",
    format="%(asctime)s %(filename)s:%(lineno)s%(funcName)20s() %(levelname)s %(message)s")
schedule_logger = logging.getLogger('schedule')
schedule_logger.setLevel(level=logging.DEBUG)

bot = TeleBot(telegram_token)

commands = [
    'Статус блокировки',
    'Разблокировать номер',
    'Заблокировать номер',
    'Замена СИМ-карты',
    'Список СИМ-карт',
    'Оплата',
    'Проверки',
    'Чья смена?',
    'Запас',
    'Прайс'
]
install_commands = [commands[8], commands[9]]
keyboard_main = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
keyboard_main.add(*[types.KeyboardButton(comm) for comm in commands])

keyboard_install = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
keyboard_install.add(*[types.KeyboardButton(comm) for comm in install_commands])

install_telegram_id = [
    int(telegram_malashin_id),
    int(telegram_sumbulov_id),
    int(telegram_my_930_id)
]


def exception_handler(func):
    """Обработчик исключений."""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as err:
            logging.critical('Error', exc_info=err)
            return None

    return wrapper

@exception_handler
def check_user(message):
    """Проверяет пользователя на право выполнения команд."""
    user_id = message.chat.id
    with sq.connect(DB) as con:
        cur = con.cursor()
        cur.execute(
            "SELECT count() FROM contacts WHERE data = ?",
            (user_id,)
        )
        result = cur.fetchone()[0]
    return result


@exception_handler
def check_number(number):
    """Проверяет номера симок на корректность."""
    number = number.strip()
    if number.isdigit():
        len_number = len(number)
        if len_number == 10:
            number = "7" + number
            return True, number
        elif len_number == 11:
            number = "7" + number[1:]
            return True, number
        else:
            logging.info(msg=f"func check_number: {number}")
            return False, number
    else:
        logging.info(msg=f"func check_number: {number}")
        return False, number


@exception_handler
def check_date(date_target):
    """Проверяет дату."""
    date_target = datetime.datetime.strptime(date_target, '%Y-%m-%d')
    return isinstance(date_target, datetime.date)


@exception_handler
def cut_msg_telegram(text_msg):
    """Обрезает сообщения по максимальному значению символов в сообщении Телеграм."""
    max_char_msg_telegram = 4096
    msgs = list()
    list_rows = text_msg.split('\n')
    one_msg = str()
    len_char = 0

    for row in list_rows:
        row += '\n'
        if len(row) + len_char > max_char_msg_telegram:
            msgs.append(one_msg)
            one_msg = row
            len_char = len(row)
        else:
            one_msg += row
            len_char += len(row)

    msgs.append(one_msg)
    return msgs


@exception_handler
def say_ok(message):
    """Отвечает согласием в чат."""
    bot.send_message(chat_id=message.chat.id, text='Ок')


@exception_handler
def mts_request_number(message):
    """Запрашивает номер симки."""
    msg = bot.send_message(chat_id=message.chat.id, text="Введи номер сим-карты")
    bot.register_next_step_handler(message=msg, callback=mts_block_info)


@exception_handler
def mts_block_info(message):
    number = message.text
    check, number = check_number(number)
    if check:
        number = Number(number=number)
        api_mts.get_block_info(number)

        if number.block is None:
            msg_text = 'Ошибка в запросе'
        else:
            msg_text = f'Добровольная блокировка: {number.block}'
            if number.block:
                msg_text += f'\nДата активации: {number.block_date.date().isoformat()}'
    else:
        msg_text = 'Неверный формат номера'
    bot.send_message(chat_id=message.chat.id, text=msg_text)


@exception_handler
def mts_del_block_num_choice(message):
    """Запрашивает когда нужно удалить блокировку на номере."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='Сейчас',
            callback_data='mts_del_block_num_now'
        ),
        types.InlineKeyboardButton(
            text='Рандом 3-12 ч.',
            callback_data='mts_del_block_num_random'
        )
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys)
    bot.send_message(
        chat_id=message.chat.id,
        text='Когда?',
        reply_markup=keyboard)


@exception_handler
def mts_del_block_request_nums(message, random_time):
    """Запрашивает список номер на удаление блокировки."""
    msg = bot.send_message(
        chat_id=message.chat.id,
        text='Введи номер или номера в столбик'
    )
    bot.register_next_step_handler(
        message=msg,
        callback=mts_del_block,
        random_time=random_time
    )


@exception_handler
def mts_del_block(message, random_time):
    """Удаляет блокировку на номере/номерах."""
    numbers = list()
    for num in message.text.split('\n'):
        check, number = check_number(num)
        if check:
            numbers.append(Number(number=number))
        else:
            text_msg = f'Ошибка: номер {num} некорректен'
            bot.send_message(message.chat.id, text=text_msg)

    if random_time:
        api_mts.turn_service_numbers_later(numbers, add_service=False)
    else:
        api_mts.turn_service_numbers(numbers, add_service=False)

    numbers.sort(key=lambda x: x.api_response.success)
    text_msg = 'Результат запроса:\n'
    for number in numbers:
        text_msg += (f'{number.number}: {number.api_response.success}, '
                     f'{number.api_response.text}\n')

    bot.send_message(message.chat.id, text=text_msg)


@exception_handler
def mts_add_block_request_nums(message):
    """Запрашивает список номер на добавление блокировки."""
    msg = bot.send_message(
        chat_id=message.chat.id,
        text='Введи номер или номера в столбик'
    )
    bot.register_next_step_handler(message=msg, callback=mts_add_block)


@exception_handler
def mts_add_block(message):
    """Добавляет блокировку на номер/номера."""
    numbers = list()
    for num in message.text.split('\n'):
        check, number = check_number(num)
        if check:
            numbers.append(Number(number=number))
        else:
            text_msg = f'Ошибка: номер {num} некорректен'
            bot.send_message(message.chat.id, text=text_msg)

    api_mts.turn_service_numbers(numbers, add_service=True)

    numbers.sort(key=lambda x: x.api_response.success)
    text_msg = 'Результат запроса:\n'
    for number in numbers:
        text_msg += (f'{number.number}: {number.api_response.success}, '
                     f'{number.api_response.text}\n')

    bot.send_message(message.chat.id, text=text_msg)


@exception_handler
def mts_exchange_choice_get_number(message):
    """Запрос на выбор способа получения номера СИМ-карты."""
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(
        types.InlineKeyboardButton(
            text='Следующий номер в очереди по дате',
            callback_data='mts_exchange_sim_next_number'
        ),
        types.InlineKeyboardButton(
            text='Ввести номер самостоятельно',
            callback_data='mts_exchange_sim_input_number'
        ),
        row_width=1
    )
    bot.send_message(
        chat_id=message.chat.id,
        text='Какой номер сим-карты?',
        reply_markup=keyboard
    )


@exception_handler
def mts_exchange_sim_input_number(message):
    """Запрос ввода номера СИМ-карты."""
    msg = bot.send_message(chat_id=message.chat.id, text='Введи номер для замены')
    bot.register_next_step_handler(message=msg, callback=mts_exchange_sim, number=True)


@exception_handler
def mts_exchange_sim(message, number=False):
    """Выводит первый номер для замены МТС или проверяет введённый и спрашивает про ICC ID."""
    if not number:
        numbers = api_dj.get_numbers_for_change()
        number, num_date = numbers[1]
        text_msg = f'Всего для замены: {len(numbers)} шт.\nСледующий: {number}\nДата: {num_date}'
        bot.send_message(chat_id=message.chat.id, text=text_msg)
    else:
        number = message.text
        check, number = check_number(number)
        if not check:
            bot.send_message(chat_id=message.chat.id, text='Неверный формат номера')
            return
    msg = bot.send_message(chat_id=message.chat.id, text='Введи ICC ID')
    bot.register_next_step_handler(message=msg, callback=mts_exchange_sim_second, number=number)


@exception_handler
def mts_exchange_sim_second(message, number):
    """Проверяет, запрашивает imsi, введённой сим-карты."""
    icc_id = message.text
    if icc_id.isdigit():
        vacant_sim_response = api_mts.get_vacant_sim_card_exchange(number, icc_id)
        if vacant_sim_response.success:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(
                types.InlineKeyboardButton(
                    text='Да',
                    callback_data=f'mts_yes_exchange_sim {number};{vacant_sim_response.imsi}'
                )
            )
            bot.send_message(
                chat_id=message.chat.id,
                text=f'{number} +> {vacant_sim_response.icc}. Меняем?',
                reply_markup=keyboard
            )
        else:
            bot.send_message(
                chat_id=message.chat.id,
                text=vacant_sim_response.text
            )


@exception_handler
def mts_yes_exchange_sim(message, call_data):
    """Отправляет номер на замену сим-карты, спрашивает про дальнейшую блокировку."""
    number, imsi = call_data.split()[1].split(';')
    api_mts.get_exchange_sim_card(number, imsi)
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(
        types.InlineKeyboardButton(
            text='Да',
            callback_data=f'mts_block_exchange_sim {number}'
        ),
        types.InlineKeyboardButton(
            text='Нет',
            callback_data='say_ok'
        ),
    )
    bot.send_message(
        chat_id=message.chat.id,
        text='Подключаем Добровольную блокировку?',
        reply_markup=keyboard
    )


@exception_handler
def mts_block_exchange_sim(message, call_data):
    number = Number(number=call_data.split()[1])
    api_mts.turn_service_numbers(number)
    if number.api_response.success:
        msg_text = f'{number}: Номер в блокировке'
    else:
        msg_text = f'{number.number}: {number.api_response.success} - {number.api_response.text}'
    bot.send_message(chat_id=message.chat.id, text=msg_text)


@exception_handler
def mts_get_account_balance():
    """Сравнивает, записывает и отправляет баланс лицевого счёта."""
    result = api_mts.get_balance()
    with sq.connect(DB) as con:
        cur = con.cursor()
        cur.execute(
            'SELECT balance FROM mts_balances ORDER BY id DESC LIMIT 1'
        )
        old_balance = cur.fetchone()[0]
        cur.execute(
            'INSERT INTO mts_balances (balance) VALUES (?)',
            (result,)
        )
    difference = float(result) - old_balance
    msg_text = f'МТС Баланс: {round(result, 2)} ({round(difference, 2)})'
    bot.send_message(chat_id=telegram_my_id, text=msg_text)


@exception_handler
def mts_check_num_balance(msg_chat_id=None):
    """Проверяет номера МТС на перерасход."""
    if msg_chat_id is None:
        msg_chat_id = telegram_my_id

    extra_money = 0
    numbers = api_mts.set_balance_numbers()
    numbers.sort(key=lambda x: x.balance, reverse=True)
    mts_normal_balances = 28, 29.75, 56, 59.5
    msg_text = 'МТС Перерасход:'
    for number in numbers:
        if number.balance not in mts_normal_balances:
            msg_text += f'\n{number.number}: {number.balance}'
            extra_money += number.balance

    msg_text +=f'\nОбщий перерасход = {extra_money}'

    msgs = cut_msg_telegram(msg_text)

    for msg in msgs:
        bot.send_message(chat_id=msg_chat_id, text=msg)


@exception_handler
def check_email(
        imap_server=imap_server_yandex,
        email_login=ya_mary_email_login,
        email_password=ya_mary_email_password,
        telegram_id=id_telegram_mary
    ):
    mailbox = imaplib.IMAP4_SSL(imap_server)
    mailbox.login(email_login, email_password)
    mailbox.select()
    unseen_msg = mailbox.uid('search', "UNSEEN", "ALL")
    id_unseen_msgs = unseen_msg[1][0].decode("utf-8").split()
    logging.info(msg=f"{email_login}: {id_unseen_msgs}")
    if id_unseen_msgs:
        bot.send_message(
            telegram_id,
            text=f"На почте {email_login} есть непрочитанные письма, в кол-ве {len(id_unseen_msgs)} шт."
        )


@exception_handler
def get_numbers_payer_or_date(message):
    """Спрашивает по дате или по плательщику прислать сим-карты."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='По ID плательщика',
            callback_data='get_numbers_id_payer'),
        types.InlineKeyboardButton(
            text='По дате',
            callback_data='get_numbers_date'),
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys)
    bot.send_message(
        message.from_user.id,
        text='Выбери',
        reply_markup=keyboard)


@exception_handler
def get_number_payer_sim_cards(message):
    """Запрашивает номер плательщика."""
    msg = bot.send_message(chat_id=message.chat.id, text="Напиши ID плательщика")
    bot.register_next_step_handler(message=msg, callback=get_list_payer_sim_cards)


@exception_handler
def get_list_payer_sim_cards(message, payer_id=None):
    """Отправляет сообщение со списком номеров сим-карт по плательщику."""
    if not payer_id:
        payer_id = message.text
    sim_cards = api_dj.get_payer_sim_cards(payer_id)
    for sim_list in sim_cards:
        bot.send_message(
            chat_id=message.chat.id,
            text='\n'.join(sim_list)
        )

@exception_handler
def get_list_date_sim_cards(message):
    """Запрашивает дату."""
    msg = bot.send_message(chat_id=message.chat.id, text='Напиши дату в формате "2000-12-31"')
    bot.register_next_step_handler(message=msg, callback=get_list_date_sim_cards_handler)


@exception_handler
def get_list_date_sim_cards_handler(message):
    """Отправляет сообщение со списком номеров сим-карт по дате."""
    date = message.text
    sim_cards = api_dj.get_date_sim_cards(date)
    for sim_list in sim_cards:
        text_msg = '\n'.join(sim_list)
        for msg_one in cut_msg_telegram(text_msg):
            bot.send_message(
                chat_id=message.chat.id,
                text=msg_one
            )


@exception_handler
def check_mts_sim_cards(msg_chat_id):
    """Сравнивает сим-карты на проекте и сайте МТС."""
    mts_id = 2
    prj_mts_sim_cards = [(num[3], num[2]) for num in api_dj.get_list_sim_cards() if num[1] == mts_id]
    site_mts_sim_cards = api_mts.get_list_all_mts_sim_cards()
    another_sim = set(prj_mts_sim_cards) ^ set(site_mts_sim_cards)
    msg_text = f'Разница {len(another_sim)} шт.'
    for num in another_sim:
        msg_text += f'\n{num[0]} {num[1]}'
    for msg in cut_msg_telegram(msg_text):
        bot.send_message(
            chat_id=msg_chat_id,
            text=msg
        )


@exception_handler
def check_active_mts_sim_cards(msg_chat_id=None):
    """Сравнивает активные сим-карты на проекте и на МТС."""
    id_chats = list()
    if msg_chat_id:
        id_chats.append(msg_chat_id)
        bot.send_message(chat_id=msg_chat_id, text='Проверка запущена. Займёт ~2 часа')
    else:
        id_chats.extend((telegram_maks_id, telegram_my_id))

    all_nums = api_mts.get_list_numbers_class()
    api_mts.get_block_info(all_nums)

    # проверяет дату заблокированных номеров,
    critical_date = datetime.datetime.now(
        tz=pytz.timezone(timezone_my)
    ) - datetime.timedelta(days=180)

    critical_date_nums = list()
    active_numbers = set()
    for number in all_nums:
        number.sim_card = None
        if number.block is None:
            continue
        elif number.block and number.block_date < critical_date:
            critical_date_nums.append(number)
        elif not number.block:
            active_numbers.add(number)

    # разблокирует и блокирует
    api_mts.turn_service_numbers(critical_date_nums, add_service=False)
    for number in critical_date_nums:
        logging.info(
            f'critical_date_del_block: {number.number}, '
            f'{number.block_date}, {number.api_response.success}, '
            f'{number.api_response.text}'
        )
        if not number.api_response.success:
            msg_text = (f'Critical_date_block-del_block: {number.number}, '
                        f'{number.api_response.success}, '
                        f'{number.api_response.text}')
            bot.send_message(chat_id=telegram_my_id, text=msg_text)
    api_mts.turn_service_numbers(critical_date_nums, add_service=True)
    for number in critical_date_nums:
        logging.info(
            f'critical_date_add_block: {number.number}, '
            f'{number.api_response.success}, {number.api_response.text}'
        )
        if not number.api_response.success:
            msg_text = (f'Critical_date_block-add_block: {number.number}, '
                        f'{number.api_response.success}, '
                        f'{number.api_response.text}')
            bot.send_message(chat_id=telegram_my_id, text=msg_text)
    logging.info(f'check_critical_block_date: finish')

    dj_all_active_mts_numbers = {
        Number(number=number.get('number')) for number
        in api_dj.api_request_all_active_mts_numbers()
    }

    sim_cards = dj_all_active_mts_numbers ^ active_numbers
    msg_text = 'Разница блокировки СИМ-карт:\n'
    msg_text += '\n'.join([num.number for num in sim_cards])
    for id_chat in id_chats:
        for msg_one in cut_msg_telegram(msg_text):
            bot.send_message(
                chat_id=id_chat,
                text=msg_one
            )


@exception_handler
def check_glonasssoft_dj_objects(msg_chat_id):
    """Выдаёт разницу объектов на CLONASSSoft."""
    date_now = datetime.datetime.today()
    glonasssoft_objs = {obj.get('imei') for obj in api_glonasssoft.request_list_objects(glonasssoft_org_id)}
    user_list_target_serv = {user.get('id') for user in api_dj.api_request_user_list() if user.get('server') == 4}
    project_objs = {
        obj.get('terminal') for obj in api_dj.api_request_object_list()
        if obj.get('wialon_user') in user_list_target_serv
        and datetime.datetime.fromisoformat(obj.get('date_change_status')) >= date_now
    }
    terminals = {term.get('id'): term.get('imei') for term in api_dj.api_request_terminal_list()}
    result = glonasssoft_objs ^ {terminals[obj] for obj in project_objs}
    msg_text = f'Разница трекеров GLONASSSoft: {len(result)}\n' + '\n'.join(result)
    bot.send_message(chat_id=msg_chat_id,  text=msg_text)


@exception_handler
def payment_request_data_payer(message):
    """Запрашивает способ определения плательщика."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='По ID',
            callback_data='payment_choice_id'
        ),
        types.InlineKeyboardButton(
            text='По сообщению',
            callback_data='payment_choice_msg'
        )
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys)
    bot.send_message(
        message.from_user.id,
        text='Выбери способ определения плательщика',
        reply_markup=keyboard)


@exception_handler
def payment_request_payer_id(message):
    """Запрашивает ID плательщика"""
    msg = bot.send_message(chat_id=message.chat.id, text='Напиши ID плательщика')
    bot.register_next_step_handler(message=msg, callback=payment_request_date)


@exception_handler
def payment_request_payer_msg(message):
    """Запрашивает сообщение плательщика у которого нужно зарегистрировать платёж в объектах."""
    msg = bot.send_message(chat_id=message.chat.id, text='Перешли сообщение плательщика')
    bot.register_next_step_handler(message=msg, callback=payment_request_payer_msg_handler)


@exception_handler
def payment_request_payer_msg_handler(message):
    """Обрабатывает сообщение плательщика."""
    try:
        telegram_id_payer = message.forward_from.id
        payer_id = api_dj.get_id_human_for_from_telegram_id(telegram_id_payer)
        if not payer_id:
            bot.send_message(chat_id=message.chat.id, text=f'{telegram_id_payer} - не зарегистрирован')
        else:
            bot.send_message(chat_id=message.chat.id, text=f'Telegram ID: {telegram_id_payer}')
        payment_request_date(message, payer_id)
    except AttributeError as _:
        bot.send_message(chat_id=message.chat.id, text=f'ID пользователя скрыт')
        payment_request_payer_id(message)


@exception_handler
def payment_request_date(message, payer_id=None):
    """Выбор даты оплаченного периода."""
    if not payer_id:
        payer_id = message.text
    date_now = datetime.date.today()
    now_month_last_day = calendar.monthrange(date_now.year, date_now.month)[1]
    date_target_now_month = date_now.strftime(f'%Y-%m-{now_month_last_day}')
    date_plus_month = date_now + relativedelta(months=1)
    date_plus_month_last_day = calendar.monthrange(date_plus_month.year, date_plus_month.month)[1]
    date_target_plus_month = date_plus_month.strftime(f'%Y-%m-{date_plus_month_last_day}')
    inline_keys = [
        types.InlineKeyboardButton(
            text=date_target_now_month,
            callback_data=f'payment_change_date {payer_id} {date_target_now_month}'
        ),
        types.InlineKeyboardButton(
            text=date_target_plus_month,
            callback_data=f'payment_change_date {payer_id} {date_target_plus_month}'
        ),
        types.InlineKeyboardButton(
            text='Произвольная дата',
            callback_data=f'payment_custom_date {payer_id}'
        ),
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys, row_width=2)
    bot.send_message(
        message.from_user.id,
        text='Выбери дату',
        reply_markup=keyboard)


@exception_handler
def payment_request_custom_date(message, call_data):
    """Запрашивает произвольную дату."""
    msg = bot.send_message(chat_id=message.chat.id, text='Введи дату в формате "2000-01-31"')
    bot.register_next_step_handler(message=msg, callback=payment_change_date, call_data=call_data)


@exception_handler
def payment_change_date(message, call_data):
    """Регистрирует оплаченную дату в объектах плательщика."""
    if len(call_data.split()) == 2:
        payer_id = call_data.split()[1]
        date_target = message.text
    else:
        payer_id, date_target = call_data.split()[1:]
    if not check_date(date_target):
        bot.send_message(message.chat.id, text='Неверный формат даты')
    else:
        api_dj.objects_change_date(payer_id, date_target)
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton(
                text='Вывести СИМ-карты плательщика',
                callback_data=f'payment_get_sim_cards_payers {payer_id}'
            )
        )
        bot.send_message(
            message.chat.id,
            text='Успех',
            reply_markup=keyboard)


@exception_handler
def request_upload_mega_exel(message):
    msg = bot.send_message(chat_id=message.chat.id, text='Загрузи МЕГАФОН')
    bot.register_next_step_handler(message=msg, callback=upload_mega_exel)


@exception_handler
def upload_mega_exel(message):
    file_id = message.document.file_id
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    excel_doc = op.load_workbook(filename=BytesIO(downloaded_file), data_only=True)
    sheet_names = excel_doc.sheetnames
    sheet = excel_doc[sheet_names[0]]
    count_row = 2

    while sheet.cell(row=count_row, column=1).value is not None:
        number = sheet.cell(row=count_row, column=1).value
        icc_id = sheet.cell(row=count_row, column=17).value
        print(number, icc_id)
        count_row += 1

    excel_doc.close()


@exception_handler
def request_upload_sim2m_exel(message):
    msg = bot.send_message(chat_id=message.chat.id, text='Загрузи СИМ2М')
    bot.register_next_step_handler(message=msg, callback=upload_sim2m_exel)


@exception_handler
def upload_sim2m_exel(message):
    file_id = message.document.file_id
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    excel_doc = op.open(filename=BytesIO(downloaded_file), data_only=True)
    sheet_names = excel_doc.sheetnames
    sheet = excel_doc[sheet_names[0]]
    count_row = 2

    while sheet.cell(row=count_row, column=1).value is not None:
        number = sheet.cell(row=count_row, column=6).value
        icc_id = sheet.cell(row=count_row, column=7).value
        status = sheet.cell(row=count_row, column=2).value
        print(number, icc_id, status)
        count_row += 1

    excel_doc.close()


def check_diff_terminals(msg_chat_id):
    """Возвращает проверку трекеров."""
    diff_trackers_all_vs_install_and_on_hands, diff_trackers_on_hands_and_obj = api_dj.get_diff_terminals()
    msg_text = 'Трекеры без объектов и не на руках:\n' + '\n'.join(diff_trackers_all_vs_install_and_on_hands)
    for msg_one in cut_msg_telegram(msg_text):
        bot.send_message(
            chat_id=msg_chat_id,
            text=msg_one
        )
    msg_text = 'Трекеры в объектах и на руках:\n' + '\n'.join(diff_trackers_on_hands_and_obj)
    for msg_one in cut_msg_telegram(msg_text):
        bot.send_message(
            chat_id=msg_chat_id,
            text=msg_one
        )


@exception_handler
def check_sim_cards_in_dj(msg_chat_id):
    """Проверка сим-карт внутри проекта."""
    sim_in_trackers_and_on_hands, sim_not_everywhere = api_dj.check_sim_cards_in_dj()
    msg_text = f'СИМ-карты и на руках и в трекерах: {len(sim_in_trackers_and_on_hands)}\n' + '\n'.join(sim_in_trackers_and_on_hands)
    for msg_one in cut_msg_telegram(msg_text):
        bot.send_message(
            chat_id=msg_chat_id,
            text=msg_one
        )
    msg_text = f'СИМ-карты не на руках и не в трекерах: {len(sim_not_everywhere)}\n' + '\n'.join(sim_not_everywhere)
    for msg_one in cut_msg_telegram(msg_text):
        bot.send_message(
            chat_id=msg_chat_id,
            text=msg_one
        )


@exception_handler
def schedule_request_date(message):
    """Запрос даты для вывода человека по графику."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='Сегодня',
            callback_data=f'schedule today'
        ),
        types.InlineKeyboardButton(
            text='Завтра',
            callback_data=f'schedule tomorrow'
        )
        # types.InlineKeyboardButton(
        #     text='Произвольная дата',
        #     callback_data=''
        # )
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys, row_width=2)
    bot.send_message(
        message.chat.id,
        text='Когда?',
        reply_markup=keyboard)


@exception_handler
def schedule_get_human(message, call_data):
    """Выводит человека чья дата смены."""
    if 'tomorrow' in call_data:
        date_target = datetime.datetime.today() + datetime.timedelta(days=1)
        date_target = date_target.date()
    else:
        date_target = datetime.date.today()
    schedule_man = api_dj.get_api_schedule_man(date_target)
    bot.send_message(
        chat_id=message.chat.id,
        text=f'{date_target} дежурит {schedule_man}'
    )


@exception_handler
def checks(message):
    """Выдаёт список проверок."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='Проверка СИМ-карт на проекте',
            callback_data='check_sim_cards_in_dj'
        ),
        types.InlineKeyboardButton(
            text='Проверка номеров',
            callback_data='check_numbers'
        ),
        types.InlineKeyboardButton(
            text='Проверка перерасхода СИМ-карт',
            callback_data='check_overspend_sim_cards'
        ),
        types.InlineKeyboardButton(
            text='Проверка блокировок СИМ-карт',
            callback_data='check_active_sim_cards'
        ),
        types.InlineKeyboardButton(
            text='Проверка объектов GLONASSsoft',
            callback_data='check_glonasssoft'
        ),
        types.InlineKeyboardButton(
            text='Потерянные трекеры',
            callback_data='check_lost_trackers'
        ),
        types.InlineKeyboardButton(
            text='Загрузить МЕГАФОН',
            callback_data='check_upload_mega_exel'
        ),
        types.InlineKeyboardButton(
            text='Загрузить СИМ2М',
            callback_data='check_upload_sim2m_exel'
        )
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys, row_width=1)
    msg_text = (
        '"Проверка СИМ-карт на проекте" - проверка наличия сим-карт в трекерах и на руках\n'
        '"Проверка номеров" - разница номеров и ICC ID на проекте и на сайте МТС\n'
        '"Проверка перерасхода СИМ-карт" - сим-карты с расходом более 28 руб.\n'
        '"Проверка блокировок СИМ-карт" - разница активных номеров на проекте и сайте МТС (~1,5 часа)\n'
        '"Проверка объектов GLONASSsoft" - разница активных объектов на проекте и GLONASSSoft\n'
        '"Потерянные трекеры" - трекеры без установок и не на руках'
    )
    bot.send_message(
        chat_id=message.chat.id,
        text=msg_text,
        reply_markup=keyboard
    )


@exception_handler
def get_stock(message):
    telegram_id = message.chat.id
    if telegram_id in install_telegram_id:
        inline_keys = [
            types.InlineKeyboardButton(
                text='Кол-во',
                callback_data=f'get_stock {telegram_id}'
            ),
            types.InlineKeyboardButton(
                text='Детально',
                callback_data=f'get_stock {telegram_id} details'
            )
        ]
        text_msg = 'Как вывести?'
    else:
        inline_keys = [
            types.InlineKeyboardButton(
                text='Малашин А.',
                callback_data=f'get_stock {telegram_malashin_id}'
            ),
            types.InlineKeyboardButton(
                text='Сумбулов А.',
                callback_data=f'get_stock {telegram_sumbulov_id}'
            ),
            types.InlineKeyboardButton(
                text='Мерзляков М.',
                callback_data=f'get_stock {telegram_maks_id}'
            ),
            types.InlineKeyboardButton(
                text='Лехтин А.',
                callback_data=f'get_stock {telegram_my_id}'
            )
        ]
        text_msg = 'Чей запас вывести?'
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys, row_width=2)
    bot.send_message(
        chat_id=telegram_id,
        text=text_msg,
        reply_markup=keyboard
    )


@exception_handler
def get_stock_handler(message, call_data):
    """Выводит кол-во и список оборудования на руках."""
    target_telegram_id = call_data.split()[1]
    result = api_dj.get_stock(target_telegram_id)
    if 'details' in call_data:
        text_result_list = list()
        for key, value in result.items():
            text_result_list.append(key)
            for imei, serial in value:
                text_result_list.append(f'{imei} {serial}')
        text_msg='\n'.join(text_result_list)
    else:
        result_len = [f'{key} - {len(value)} шт.' for key, value in result.items()]
        text_msg = '\n'.join(result_len)
    bot.send_message(
        chat_id=message.chat.id,
        text=text_msg
    )


@exception_handler
def get_price(message):
    """Запрашивает тип прайса для вывода."""
    inline_keys = [
        types.InlineKeyboardButton(
            text='Выезды',
            callback_data='get_price logistics'
        ),
        types.InlineKeyboardButton(
            text='Оборудование',
            callback_data='get_price trackers'
        ),
        types.InlineKeyboardButton(
            text='Услуги',
            callback_data='get_price services'
        )
    ]
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(*inline_keys)
    bot.send_message(
        message.from_user.id,
        text='Выбери',
        reply_markup=keyboard
    )


@exception_handler
def get_price_type(message, call_data):
    """Отправляет сообщение с прайсом."""
    target_price = call_data.split()[1]
    result = str()
    match target_price:
        case 'logistics':
            result = api_dj.get_price_logistic()
        case 'trackers':
            result = api_dj.get_price_trackers()
        case 'services':
            result = api_dj.get_services()
    bot.send_message(
        message.chat.id,
        text='\n'.join(result)
    )


@exception_handler
def morning_check():
    """
    Утренний скрипт:
    1. проверка баланса ЛС,
    2. проверка перерасхода номеров,
    3. проверка наличие сим-карт на проекте и МТС.
    """
    logging.info(msg='Start morning_check')
    mts_get_account_balance()
    check_sim_cards_in_dj(msg_chat_id=telegram_job_id)
    mts_check_num_balance()
    check_mts_sim_cards(msg_chat_id=telegram_job_id)
    check_diff_terminals(msg_chat_id=telegram_job_id)
    check_glonasssoft_dj_objects(msg_chat_id=telegram_job_id)


@exception_handler
def schedule_main():
        schedule.every().day.at('06:30', timezone(timezone_my)).do(
            check_active_mts_sim_cards
        )
        schedule.every().day.at('08:30', timezone(timezone_my)).do(
            morning_check
        )
        schedule.every().day.at('09:00', timezone(timezone_my)).do(check_email)
        schedule.every().day.at('15:00', timezone(timezone_my)).do(check_email)
        schedule.every().day.at('21:00', timezone(timezone_my)).do(check_email)

        while True:
            schedule.run_pending()
            time.sleep(1)


@bot.message_handler(content_types=['photo'], func=lambda message: message.chat.type in ['group', 'supergroup'])
def handler_group_photo(message):
    file_info = bot.get_file(message.photo[-1].file_id)
    downloaded_file = Image.open(BytesIO(bot.download_file(file_info.file_path)))
    decoded = decode(downloaded_file)
    result = [code.data.decode("utf-8") for code in decoded]
    if result:
        bot.send_message(
            message.chat.id,
            text='\n'.join(result),
            business_connection_id=message.business_connection_id
        )


@bot.message_handler(commands=['start'])
def start_message(message):
    if check_user(message):
        text = "Привет! Лови клавиатуру!\nЧтобы получить описание введи /help"
        bot.send_message(
            message.chat.id,
            text=text,
            reply_markup=keyboard_main
        )
    else:
        bot.send_message(message.chat.id, text="Привет!")


@bot.message_handler(commands=['commands', 'help'])
def help_message(message):
    if check_user(message):
        text = (
            '1. Можно вводить команды, не дожидаясь завершения другой\n'
            '2. Можно вводить номера через 7, 8 или без первой цифры, т.е. 79998887766, 89998887766, 9998887766\n'
            '3. Можно вводить команду с несколькими номерами, '
            'нужно вводить их в столбик, без дополнительных знаков. Например:\n'
            '9998887766\n'
            '79998887766\n'
            '89998887766\n'
            '4. "Разблокировка рандом" разблокирует номера в диапазоне 3-12 часов после отправки команды\n'
        )
        bot.send_message(
            message.chat.id,
            text=text,
            reply_markup=keyboard_main
        )
    elif message.chat.id in install_telegram_id:
        bot.send_message(
            message.chat.id,
            text='Лови клавиатуру',
            reply_markup=keyboard_install
        )
    else:
        bot.send_message(chat_id=message.chat.id, text='В другой раз')


@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    if 'get_numbers_date' in call.data:
        get_list_date_sim_cards(call.message)
    elif 'get_numbers_id_payer' in call.data:
        get_number_payer_sim_cards(call.message)

    elif 'get_stock' in call.data:
        get_stock_handler(call.message, call.data)

    elif 'check_numbers' in call.data:
        check_mts_sim_cards(call.message.chat.id)
    elif 'check_overspend_sim_cards' in call.data:
        mts_check_num_balance(msg_chat_id=call.message.chat.id)
    elif 'check_active_sim_cards' in call.data:
        check_active_mts_sim_cards(call.message.chat.id)
    elif 'check_glonasssoft' in call.data:
        check_glonasssoft_dj_objects(call.message.chat.id)
    elif 'check_lost_trackers' in call.data:
        check_diff_terminals(call.message.chat.id)
    elif 'check_sim_cards_in_dj' in call.data:
        check_sim_cards_in_dj(call.message.chat.id)
    elif 'check_upload_mega_exel' in call.data:
        request_upload_mega_exel(call.message)
    elif 'check_upload_sim2m_exel' in call.data:
        request_upload_sim2m_exel(call.message)

    elif 'mts_del_block_num_now' in call.data:
        mts_del_block_request_nums(call.message, random_time=False)
    elif 'mts_del_block_num_random' in call.data:
        mts_del_block_request_nums(call.message, random_time=True)

    elif 'mts_exchange_sim_next_number' in call.data:
        mts_exchange_sim(call.message)
    elif 'mts_exchange_sim_input_number' in call.data:
        mts_exchange_sim_input_number(call.message)
    elif 'mts_yes_exchange_sim' in call.data:
        mts_yes_exchange_sim(call.message, call.data)
    elif 'mts_block_exchange_sim' in call.data:
        mts_block_exchange_sim(call.message, call.data)

    elif 'payment_choice_id' in call.data:
        payment_request_payer_id(call.message)
    elif 'payment_choice_msg' in call.data:
        payment_request_payer_msg(call.message)
    elif 'payment_change_date' in call.data:
        payment_change_date(call.message, call.data)
    elif 'payment_custom_date' in call.data:
        payment_request_custom_date(call.message, call.data)
    elif 'payment_get_sim_cards_payers' in call.data:
        get_list_payer_sim_cards(call.message, payer_id=call.data.split()[1])

    elif 'get_price' in call.data:
        get_price_type(call.message, call.data)

    elif 'schedule' in call.data:
        schedule_get_human(call.message, call.data)

    elif 'say_ok' in call.data:
        say_ok(call.message)


@bot.message_handler(content_types=['text'])
def take_text(message):
    if check_user(message) or message.chat.id in install_telegram_id:
        if message.text.lower() == commands[0].lower():
            mts_request_number(message)
        elif message.text.lower() == commands[1].lower():
            mts_del_block_num_choice(message)
        elif message.text.lower() == commands[2].lower():
            mts_add_block_request_nums(message)
        elif message.text.lower() == commands[3].lower():
            mts_exchange_choice_get_number(message)
        elif message.text.lower() == commands[4].lower():
            get_numbers_payer_or_date(message)
        elif message.text.lower() == commands[5].lower():
            payment_request_data_payer(message)
        elif message.text.lower() == commands[6].lower():
            checks(message)
        elif message.text.lower() == commands[7].lower():
            schedule_request_date(message)
        elif message.text.lower() == commands[8].lower():
            get_stock(message)
        elif message.text.lower() == commands[9].lower():
            get_price(message)
        else:
            logging.warning(f'func take_text: not understand question: {message.text}')
            bot.send_message(message.chat.id, 'Я не понимаю, к сожалению')


if __name__ == '__main__':
    threading.Thread(target=schedule_main).start()
    bot.infinity_polling()
